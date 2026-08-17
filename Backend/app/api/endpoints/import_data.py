from fastapi import (
    APIRouter,
    Depends,
    UploadFile,
    File,
    HTTPException,
    Form,
)
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError
from sqlalchemy import select, update, or_
import pandas as pd
import io
import logging
import re
from datetime import datetime
from typing import Optional, Dict, Any, List, cast

from app.db.session import get_db
from app.api.deps import get_current_user
from app.models.inventory import Item, Activo, StockBulk
from app.models.business import Cliente, Proyecto, Proveedor
from app.models.guarantees import Garantia
from app.crud.crud_audit import create_audit_log

router = APIRouter()
logger = logging.getLogger(__name__)


# Funciones auxiliares para limpieza y transformación de datos


def _clean(val, default=None):
    if val is None:
        return default
    try:
        import math
        if math.isnan(float(val)):
            return default
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    return s if s and s.lower() not in ("nan", "none", "") else default


def _to_float(val, default=0.0):
    try:
        result = float(val)
        import math
        return default if math.isnan(result) or math.isinf(result) else result
    except (TypeError, ValueError):
        return default


def _to_date(val):
    if val is None or pd.isna(val):
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    try:
        parsed = pd.to_datetime(val)
        return None if pd.isna(parsed) else parsed.date()
    except Exception:
        return None


def _categoria_from_sheet(sheet_name) -> str:
    s = str(sheet_name).upper()
    if "HERRAMIENTA" in s:
        return "HERRAMIENTA_LAB"
    if "MONITOREO" in s:
        return "MONITOREO"
    if "MANTENIMIENTO" in s:
        return "MANTENIMIENTO"
    if "INSTALACION" in s or "ESTANTE" in s or "ESCRITORIO" in s:
        return "INSTALACION"
    if "SOLUCIONES" in s:
        return "SOLUCIONES"
    if "CONSUMIBLE" in s or "ELKIN" in s or "CAJA" in s:
        return "CONSUMIBLE"
    return "INSTALACION"


def _slug(s: str) -> Optional[str]:
    """Convierte texto a slug en MAYÚSCULAS, máximo 50 chars (deja espacio para sufijos)."""
    if not s:
        return None
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    s = s.strip("_")[:50]
    return s or None



# UPSERT: crea o actualiza un Item y su StockBulk


async def _upsert_item_and_stock(
    db: AsyncSession,
    nombre: str,
    referencia: Optional[str],
    codigo: Optional[str],
    marca: Optional[str],
    categoria: str,
    sub_categoria: Optional[str],
    cantidad: float,
    stock_min: int,
    compra_max: int,
    always_create_new: bool = False,
    modo_stock: str = "sumar",
) -> Item:
    """
    Crea un Item nuevo o reutiliza uno existente.
    - Si always_create_new=False -> busca por referencia/codigo/nombre y si existe, solo actualiza stock.
    - Si always_create_new=True -> SIEMPRE crea uno nuevo, agregando sufijos si hay duplicados.
    - modo_stock determina cómo se actualiza el stock del item existente:
      - "sumar" -> el stock del Excel se SUMA al stock actual (nunca baja).
      - "reemplazar" -> el stock queda igual al valor del Excel.
    """

    ref = _clean(referencia)
    cod = _clean(codigo)

    if not ref and nombre:
        ref = _slug(nombre)
    if not cod and not ref and nombre:
        cod = _slug(nombre)

    # SIEMPRE verificar duplicados de referencia (incluso cuando always_create_new=True)
    if always_create_new and ref:
        base_ref = ref[:40]
        candidate_ref = base_ref
        suffix = 1
        while True:
            try:
                res = await db.execute(
                    select(Item).where(Item.referencia == candidate_ref)
                )
                if not res.scalars().first():
                    break
            except Exception:
                break

            candidate_ref = f"{base_ref}_{suffix}"
            suffix += 1
            if suffix > 999:
                candidate_ref = f"{base_ref[:30]}_{int(datetime.utcnow().timestamp())}"
                break
        ref = candidate_ref

    # Si NO es always_create_new, buscar uno existente
    if not always_create_new:
        result = await db.execute(
            select(Item).where(
                (Item.referencia == ref) |
                (Item.codigo_item_interno == cod) |
                (Item.nombre_equipo == nombre)
            )
        )
        existing = result.scalars().first()
        if existing:
            result_stock = await db.execute(
                select(StockBulk).where(StockBulk.id_item == existing.id_item)
            )
            stock_row = result_stock.scalars().first()
            if stock_row:
                if modo_stock == "sumar":
                    stock_row.cantidad_actual = (
                        float(stock_row.cantidad_actual or 0) + cantidad
                    )
                else:
                    stock_row.cantidad_actual = cantidad
            else:
                db.add(
                    StockBulk(id_item=existing.id_item, cantidad_actual=cantidad)
                )
            return existing

    # Generar codigo_item_interno único
    generated_code = (
        cod
        or _slug(nombre)
        or f"ITEM_{int(datetime.utcnow().timestamp())}"
    )
    base_code = generated_code[:40]
    candidate_code = base_code
    suffix = 1
    while True:
        try:
            res = await db.execute(
                select(Item).where(Item.codigo_item_interno == candidate_code)
            )
            if not res.scalars().first():
                break
        except Exception:
            break

        candidate_code = f"{base_code}_{suffix}"
        suffix += 1
        if suffix > 999:
            candidate_code = f"{base_code[:30]}_{int(datetime.utcnow().timestamp())}"
            break

    item = Item(
        nombre_equipo=nombre or "N/A",
        referencia=ref or candidate_code,
        codigo_item_interno=candidate_code,
        marca=marca,
        categoria=categoria,
        sub_categoria=sub_categoria,
        stock_minimo=stock_min,
        compra_maxima=compra_max,
    )
    db.add(item)
    await db.flush()
    db.add(StockBulk(id_item=item.id_item, cantidad_actual=cantidad))
    return item


# ============================================================
# Procesadores por archivo
# ============================================================

async def _procesar_inventario_laboratorio(
    xl: pd.ExcelFile,
    db: AsyncSession,
    id_proyecto: Optional[int] = None,
    id_cliente: Optional[int] = None,
    modo_stock: str = "sumar",
):
    """
    Procesa Inventario_laboratorio.xlsx
    """
    items_creados = 0
    activos_creados = 0

    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        desc_col = next(
            (c for c in df.columns if "DESCRIPCION" in c.upper() or "DESCRIPCIÓN" in c.upper()),
            None,
        )
        modelo_col = next(
            (c for c in df.columns if c.upper() in ("MODELO", "REFERENCIA")),
            None,
        )
        marca_col = next(
            (c for c in df.columns if "MARCA" in c.upper()),
            None,
        )
        cantidad_col = next(
            (c for c in df.columns if "CANTIDAD" in c.upper()),
            None,
        )
        ubicacion_col = next(
            (c for c in df.columns if "UBICACI" in c.upper()),
            None,
        )
        estado_col = next(
            (c for c in df.columns if c.upper() == "ESTADO"),
            None,
        )
        codigo_col = next(
            (c for c in df.columns if c.upper() == "CODIGO"),
            None,
        )
        activo_fijo_col = next(
            (c for c in df.columns if "ACTIVO" in c.upper() and "FIJO" in c.upper()),
            None,
        )

        if not desc_col:
            logger.warning(f"Hoja '{sheet_name}' sin columna de descripción, omitida.")
            continue

        categoria = _categoria_from_sheet(sheet_name)

        for _, row in df.iterrows():
            nombre = _clean(row.get(desc_col))
            if not nombre:
                continue

            modelo = _clean(row.get(modelo_col)) if modelo_col else None
            marca = _clean(row.get(marca_col)) if marca_col else None
            cantidad = _to_float(row.get(cantidad_col, 0) if cantidad_col else 0)
            ubicacion = _clean(row.get(ubicacion_col)) if ubicacion_col else None
            codigo = _clean(row.get(codigo_col)) if codigo_col else None
            activo_fijo = _clean(row.get(activo_fijo_col)) if activo_fijo_col else None

            # Estado / condición física del activo
            estado_raw = _clean(row.get(estado_col)) if estado_col else None
            condicion_map = {
                "NUEVO": "NUEVO",
                "BUEN ESTADO": "USADO_BUENO",
                "USADO BUEN ESTADO": "USADO_BUENO",
                "BUENO": "USADO_BUENO",
                "PARA REPARAR": "PARA_REPARAR",
                "SULFATADO": "SULFATADO",
                "DAÑADO": "DAÑADO",
                "DESMONTE": "USADO_BUENO",
            }
            condicion = "NUEVO"
            if estado_raw:
                condicion = condicion_map.get(estado_raw.upper().strip(), "NUEVO")

            item = await _upsert_item_and_stock(
                db=db,
                nombre=nombre,
                referencia=modelo,
                codigo=codigo,
                marca=marca,
                categoria=categoria,
                sub_categoria=str(sheet_name).strip(),
                cantidad=cantidad,
                stock_min=2,
                compra_max=10,
                modo_stock=modo_stock,
            )
            items_creados += 1

            # Crear Activo con serial sintético
            if ubicacion:
                serial_base = f"{(modelo or nombre)[:15].replace(' ','_').upper()}-LAB"
                serial_final = f"{serial_base}-{activo_fijo or 'SIN_AF'}"

                res = await db.execute(
                    select(Activo).where(Activo.serial == serial_final)
                )
                if not res.scalars().first():
                    observaciones = f"Importado desde hoja: {sheet_name}"
                    if id_cliente is not None:
                        observaciones += f" | Cliente ID: {id_cliente}"

                    activo = Activo(
                        id_item=item.id_item,
                        serial=serial_final,
                        estado_actual="LABORATORIO",
                        condicion_fisica=condicion,
                        ubicacion_fisica=ubicacion,
                        activo_fijo_securitas=activo_fijo,
                        area_asignada="LABORATORIO",
                        id_proyecto_actual=id_proyecto,
                        observaciones=observaciones,
                    )
                    db.add(activo)
                    activos_creados += 1

        await db.flush()
        logger.info(f"Hoja '{sheet_name}': procesada.")

    return {"items": items_creados, "activos": activos_creados}


async def _crear_cliente_si_no_existe(db: AsyncSession, nombre: str) -> Optional[int]:
    if not nombre:
        return None
    nombre = nombre.strip()
    result = await db.execute(select(Cliente).where(Cliente.nombre == nombre))
    existing = result.scalars().first()
    if existing:
        return cast(int, existing.id_cliente)
    cliente = Cliente(nombre=nombre, tipo_cliente="CORPORATIVO")
    db.add(cliente)
    await db.flush()
    await db.refresh(cliente)
    logger.info(f"  Cliente creado: {nombre}")
    return cast(int, cliente.id_cliente)

async def _procesar_inventario_clientes(
    xl: pd.ExcelFile, db: AsyncSession, modo_stock: str = "sumar"
):
    """
    Procesa Formato_Inventario_Clientes_*.xlsx
    """
    HOJAS_DATOS = [
        "Inventario Procafecol",
        "Inventario ISIMO",
        "Inventario Alsea",
        "Inventario Arcos Dorados",
        "Inventario Consolidado",
    ]
    items_creados = 0
    clientes_creados = 0
    clientes_vistos = set()

    for sheet_name in xl.sheet_names:
        if sheet_name not in HOJAS_DATOS:
            continue

        df = xl.parse(sheet_name, header=1)
        df.columns = [str(c).strip() for c in df.columns]

        def find_col(keywords):
            cols_upper = [c.upper() for c in df.columns]
            for kw in keywords:
                kw_up = kw.upper()
                for i, c in enumerate(cols_upper):
                    if c == kw_up:
                        return df.columns[i]
            for kw in keywords:
                kw_up = kw.upper()
                for i, c in enumerate(cols_upper):
                    if kw_up in c:
                        return df.columns[i]
            return None

        codigo_col = find_col(["CÓDIGO", "CODIGO"])
        item_col = find_col(["ITEM"])
        marca_col = find_col(["MARCA"])
        ref_col = find_col(["REFERENCIA"])
        stock_col = find_col(["STOCK ACTUAL", "STOCK"])
        recompra_col = find_col(["PUNTO DE RECOMPRA"])
        compra_col = find_col(["COMPRA MÁXIMA", "COMPRA MAXIMA"])
        cliente_col = find_col(["CORPORATIVO"])
        sistema_col = find_col(["SISTEMA"])
        ceco_col = find_col(["CECO COMPRA", "SECO COMPRA", "CECO"])

        if not item_col:
            logger.warning(f"Hoja '{sheet_name}' sin columna ITEM, omitida.")
            continue

        for _, row in df.iterrows():
            nombre = _clean(row.get(item_col))
            if not nombre or nombre.upper() in ("ITEM", "NAN"):
                continue

            # Crear cliente si es nuevo
            cliente_nombre = _clean(row.get(cliente_col)) if cliente_col else None
            if cliente_nombre and cliente_nombre not in clientes_vistos:
                clientes_vistos.add(cliente_nombre)
                cid = await _crear_cliente_si_no_existe(db, cliente_nombre)
                if cid:
                    clientes_creados += 1

            codigo = _clean(row.get(codigo_col)) if codigo_col else None
            if codigo:
                try:
                    codigo = str(int(float(codigo)))
                except ValueError:
                    pass

            referencia = _clean(row.get(ref_col)) if ref_col else None
            if referencia:
                referencia = referencia.strip().lstrip("\t")

            marca = _clean(row.get(marca_col)) if marca_col else None
            stock_actual = _to_float(row.get(stock_col, 0) if stock_col else 0)
            punto_recompra = _to_float(row.get(recompra_col, 0) if recompra_col else 0)
            compra_max = int(_to_float(row.get(compra_col, 20) if compra_col else 20)) or 20

            sistema = _clean(row.get(sistema_col)) if sistema_col else None
            cat_map = {
                "CCTV": "MONITOREO",
                "ALARMA": "MONITOREO",
                "MANTENIMIENTO": "MANTENIMIENTO",
                "INSTALACION": "INSTALACION",
            }
            categoria = cat_map.get((sistema or "").upper(), "MONITOREO")

            await _upsert_item_and_stock(
                db=db,
                nombre=nombre,
                referencia=referencia,
                codigo=codigo,
                marca=marca,
                categoria=categoria,
                sub_categoria=sistema,
                cantidad=stock_actual,
                stock_min=int(punto_recompra) if punto_recompra else 5,
                compra_max=compra_max,
                always_create_new=False,
                modo_stock=modo_stock,
            )
            items_creados += 1

        await db.flush()
        logger.info(f"Hoja '{sheet_name}': procesada.")

    return {"items": items_creados, "clientes": clientes_creados}


async def _procesar_garantias(
    xl: pd.ExcelFile, db: AsyncSession, modo_stock: str = "sumar"
):
    """
    Procesa ASIGNACION_NUMERO_DE_CASO_*.xlsx
    """
    garantias_creadas = 0
    items_stock_creados = 0
    clientes_creados = 0
    proyectos_creados = 0
    clientes_cache = {}
    proyectos_cache = {}

    # Hoja GARANTIAS
    if "GARANTIAS" in xl.sheet_names:
        try:
            df = xl.parse("GARANTIAS", header=1)
        except (ValueError, StopIteration, IndexError) as e:
            logger.warning(
                f"Hoja GARANTIAS sin filas de datos suficientes: {e}"
            )
            df = None
        if df is None or len(df) == 0:
            raise HTTPException(
                status_code=400,
                detail="El archivo no contiene filas de datos en la hoja GARANTIAS. Descargue la plantilla, llénela y súbala nuevamente.",
            )
        df.columns = [str(c).strip() for c in df.columns]

        ESTADO_MAP = {
            "CASO FINALIZADO": "ENTREGADO_CLIENTE",
            "ENVIADO A PROVEEDOR": "ENVIADO_PROVEEDOR",
            "RECIBIDO DE PROVEEDOR": "RECIBIDO_PROVEEDOR",
            "EN PROCESO": "ENVIADO_PROVEEDOR",
            "RESUELTO": "RESUELTO_REEMPLAZADO",
        }

        def find_col(df, *keywords):
            for kw in keywords:
                for c in df.columns:
                    if kw.upper() in c.upper():
                        return c
            return None

        caso_col = find_col(df, "NUMERO DE CASO")
        serial_col = find_col(df, "SERIAL")
        ref_col = find_col(df, "REFERENCIA DE EQUIPO")
        desc_col = find_col(df, "DESCRIPCIÓN DE EQUIPO", "DESCRIPCION")
        falla_col = find_col(df, "FALLA")
        rma_col = find_col(df, "RMA")
        factura_col = find_col(df, "NUMERO DE FACTURA")
        envio_col = find_col(df, "FECHA DE ENVIO")
        recibido_col = find_col(df, "FECHA DE  RECIBIDO", "FECHA RECIBIDO", "RECIBIDO DE EQUIPO")
        estado_col = find_col(df, "ESTADO DEL CASO")
        comentario_col = find_col(df, "COMENTARIOS DEL PROCESO")
        credencial_col = find_col(df, "CREDENCIALES", "IP/CLAVES", "CONTRASEÑAS", "OBSERVACIONES")
        area_col = find_col(df, "AREA", "ÁREA")
        meses_col = find_col(df, "MESES")
        cliente_col = find_col(df, "CLIENTE")
        proyecto_col = find_col(df, "PROYECTO")

        for _, row in df.iterrows():
            caso = _clean(row.get(caso_col)) if caso_col else None
            if not caso or caso.upper() in ("NUMERO DE CASO", "NAN"):
                continue

            # Crear Cliente si no existe
            cliente_nombre = _clean(row.get(cliente_col)) if cliente_col else None
            id_cliente = None
            if cliente_nombre and cliente_nombre not in clientes_cache:
                cid = await _crear_cliente_si_no_existe(db, cliente_nombre)
                if cid:
                    clientes_cache[cliente_nombre] = cid
                    clientes_creados += 1
            id_cliente = clientes_cache.get(cliente_nombre)

            # Crear Proyecto si no existe
            proyecto_nombre = _clean(row.get(proyecto_col)) if proyecto_col else None
            id_proyecto = None
            if proyecto_nombre:
                cache_key = f"{cliente_nombre}|{proyecto_nombre}"
                if cache_key not in proyectos_cache:
                    result = await db.execute(
                        select(Proyecto).where(
                            Proyecto.nombre_proyecto == proyecto_nombre,
                            Proyecto.id_cliente == id_cliente,
                        )
                    )
                    existing_p = result.scalars().first()
                    if existing_p:
                        proyectos_cache[cache_key] = existing_p.id_proyecto
                    else:
                        p = Proyecto(
                            nombre_proyecto=proyecto_nombre,
                            id_cliente=id_cliente,
                            estado="ACTIVO",
                        )
                        db.add(p)
                        await db.flush()
                        await db.refresh(p)
                        proyectos_cache[cache_key] = p.id_proyecto
                        proyectos_creados += 1
                id_proyecto = proyectos_cache.get(cache_key)

            serial = _clean(row.get(serial_col)) if serial_col else None
            if not serial:
                continue

            res = await db.execute(select(Activo).where(Activo.serial == serial))
            activo = res.scalars().first()

            if not activo:
                nombre = _clean(row.get(desc_col)) if desc_col else "EQUIPO GARANTIA"
                referencia = _clean(row.get(ref_col)) if ref_col else None
                item = await _upsert_item_and_stock(
                    db=db,
                    nombre=nombre or "EQUIPO GARANTIA",
                    referencia=referencia,
                    codigo=None,
                    marca=None,
                    categoria="MONITOREO",
                    sub_categoria="GARANTIA",
                    cantidad=0,
                    stock_min=1,
                    compra_max=5,
                    modo_stock=modo_stock,
                )
                activo = Activo(
                    id_item=item.id_item,
                    serial=serial,
                    estado_actual="EN_GARANTIA",
                    condicion_fisica="PARA_REPARAR",
                    area_asignada="LABORATORIO",
                    observaciones=f"Creado desde garantía {caso}",
                    id_cliente_actual=id_cliente,
                    id_proyecto_actual=id_proyecto,
                )
                db.add(activo)
                await db.flush()
            else:
                setattr(activo, "estado_actual", "EN_GARANTIA")

            estado_raw = _clean(row.get(estado_col)) if estado_col else None
            estado_garantia = "REGISTRADO"
            if estado_raw:
                for k, v in ESTADO_MAP.items():
                    if k in estado_raw.upper():
                        estado_garantia = v
                        break

            res2 = await db.execute(
                select(Garantia).where(Garantia.numero_caso_interno == caso)
            )
            if res2.scalars().first():
                continue

            garantia = Garantia(
                id_activo=activo.id_activo,
                numero_caso_interno=caso,
                rma_proveedor=_clean(row.get(rma_col)) if rma_col else None,
                numero_factura_compra=_clean(row.get(factura_col)) if factura_col else None,
                fecha_envio=_to_date(row.get(envio_col)) if envio_col else None,
                fecha_recibido_reparado=_to_date(row.get(recibido_col)) if recibido_col else None,
                credenciales_equipo=_clean(row.get(credencial_col)) if credencial_col else None,
                area_origen=_clean(row.get(area_col)) if area_col else None,
                meses_garantia=int(_to_float(row.get(meses_col, 0))) if meses_col else None,
                falla_reportada=_clean(row.get(falla_col)) if falla_col else None,
                comentarios_proceso=_clean(row.get(comentario_col)) if comentario_col else None,
                estado_proceso=estado_garantia,
            )
            db.add(garantia)
            garantias_creadas += 1

        await db.flush()

    # Hojas STOCK
    HOJAS_STOCK = [
        "STOCK MONITOREO",
        "STOCK MANTENIMIENTO",
        "STOCK INSTALACION",
        "STOCK SOLUCIONES",
    ]
    AREA_MAP = {
        "STOCK MONITOREO": "MONITOREO",
        "STOCK MANTENIMIENTO": "MANTENIMIENTO",
        "STOCK INSTALACION": "INSTALACION",
        "STOCK SOLUCIONES": "SOLUCIONES",
    }

    for sheet_name in HOJAS_STOCK:
        if sheet_name not in xl.sheet_names:
            continue
        df = xl.parse(sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        for _, row in df.iterrows():
            nombre = _clean(row.get("DESCRIPCIÓN DE EQUIPO"))
            if not nombre:
                continue
            referencia = _clean(row.get("REFERENCIA DE EQUIPO"))
            categoria_map = {
                "STOCK MONITOREO": "MONITOREO",
                "STOCK MANTENIMIENTO": "MANTENIMIENTO",
                "STOCK INSTALACION": "INSTALACION",
                "STOCK SOLUCIONES": "SOLUCIONES",
            }
            await _upsert_item_and_stock(
                db=db,
                nombre=nombre,
                referencia=referencia,
                codigo=None,
                marca=None,
                categoria=categoria_map.get(sheet_name, "MONITOREO"),
                sub_categoria="STOCK LABORATORIO",
                cantidad=1,
                stock_min=1,
                compra_max=5,
                modo_stock=modo_stock,
            )
            items_stock_creados += 1

        await db.flush()
        logger.info(f"Hoja '{sheet_name}': procesada.")

    return {"garantias": garantias_creadas, "items_stock": items_stock_creados, "clientes": clientes_creados, "proyectos": proyectos_creados}


async def _procesar_desmontes(
    xl: pd.ExcelFile,
    db: AsyncSession,
    id_proyecto: Optional[int] = None,
    id_cliente: Optional[int] = None,
    modo_stock: str = "sumar",
):
    """
    Procesa Plantilla_Desmontes.xlsx (hoja DESMONTES).
    Crea items, stock y activos desmontados que ingresan al laboratorio.
    """
    items_creados = 0
    activos_creados = 0
    clientes_creados = 0
    proyectos_creados = 0
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    from app.models.inventory import MovimientoInventario

    sheet_name = next(
        (s for s in xl.sheet_names if "DESMONTE" in str(s).upper()), None
    )
    if not sheet_name:
        logger.warning("Plantilla de desmontes sin hoja 'DESMONTES'.")
        return {"items": 0, "activos": 0, "clientes": 0, "proyectos": 0}

    df = xl.parse(sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    desc_col = next(
        (c for c in df.columns if "DESCRIPCION" in c.upper() or "DESCRIPCIÓN" in c.upper()),
        None,
    )
    modelo_col = next(
        (c for c in df.columns if c.upper() in ("MODELO", "REFERENCIA")),
        None,
    )
    marca_col = next((c for c in df.columns if "MARCA" in c.upper()), None)
    cantidad_col = next((c for c in df.columns if "CANTIDAD" in c.upper()), None)
    ubicacion_col = next((c for c in df.columns if "UBICACI" in c.upper()), None)
    serial_col = next((c for c in df.columns if c.upper() == "SERIAL"), None)
    activo_fijo_col = next(
        (c for c in df.columns if "ACTIVO" in c.upper() and "FIJO" in c.upper()),
        None,
    )
    estado_col = next((c for c in df.columns if c.upper() == "ESTADO"), None)
    obs_col = next(
        (c for c in df.columns if "OBSERVACION" in c.upper()), None
    )
    cliente_col = next((c for c in df.columns if c.upper() == "CLIENTE"), None)
    proyecto_col = next((c for c in df.columns if c.upper() == "PROYECTO"), None)

    if not desc_col:
        logger.warning(f"Hoja '{sheet_name}' sin columna de descripción.")
        return {"items": 0, "activos": 0, "clientes": 0, "proyectos": 0}

    condicion_map = {
        "NUEVO": "NUEVO",
        "BUEN ESTADO": "USADO_BUENO",
        "USADO BUEN ESTADO": "USADO_BUENO",
        "BUENO": "USADO_BUENO",
        "PARA REPARAR": "PARA_REPARAR",
        "SULFATADO": "SULFATADO",
        "DAÑADO": "DAÑADO",
        "DANADO": "DAÑADO",
        "DESMONTE": "USADO_BUENO",
    }

    for _, row in df.iterrows():
        nombre = _clean(row.get(desc_col))
        if not nombre:
            continue

        marca = _clean(row.get(marca_col)) if marca_col else None
        modelo = _clean(row.get(modelo_col)) if modelo_col else None
        cantidad = int(_to_float(row.get(cantidad_col, 0) if cantidad_col else 0))
        if cantidad < 1:
            cantidad = 1
        ubicacion = _clean(row.get(ubicacion_col)) if ubicacion_col else None
        serial_row = _clean(row.get(serial_col)) if serial_col else None
        activo_fijo = _clean(row.get(activo_fijo_col)) if activo_fijo_col else None
        obs = _clean(row.get(obs_col)) if obs_col else None
        estado_raw = _clean(row.get(estado_col)) if estado_col else None

        condicion = "USADO_BUENO"
        if estado_raw:
            condicion = condicion_map.get(estado_raw.upper().strip(), "USADO_BUENO")

        # Resolver cliente/proyecto del parámetro o de las columnas
        cliente_id = id_cliente
        proyecto_id = id_proyecto
        cliente_nombre = _clean(row.get(cliente_col)) if cliente_col else None
        if cliente_nombre and not cliente_id:
            cliente_id = await _crear_cliente_si_no_existe(db, cliente_nombre)
            if cliente_id:
                clientes_creados += 1
        proyecto_nombre = _clean(row.get(proyecto_col)) if proyecto_col else None
        if proyecto_nombre and not proyecto_id:
            result = await db.execute(
                select(Proyecto).where(
                    Proyecto.nombre_proyecto == proyecto_nombre,
                    Proyecto.id_cliente == cliente_id,
                )
            )
            existing_p = result.scalars().first()
            if existing_p:
                proyecto_id = existing_p.id_proyecto
            else:
                p = Proyecto(
                    nombre_proyecto=proyecto_nombre,
                    id_cliente=cliente_id,
                    estado="ACTIVO",
                )
                db.add(p)
                await db.flush()
                await db.refresh(p)
                proyecto_id = p.id_proyecto
                proyectos_creados += 1

        item = await _upsert_item_and_stock(
            db=db,
            nombre=nombre,
            referencia=modelo,
            codigo=None,
            marca=marca,
            categoria="INSTALACION",
            sub_categoria="DESMONTE",
            cantidad=cantidad,
            stock_min=2,
            compra_max=10,
            modo_stock=modo_stock,
        )
        items_creados += 1

        # Crear activos (uno por unidad)
        seriales_creados = []
        for i in range(cantidad):
            if serial_row and i == 0:
                serial = serial_row
            elif serial_row:
                serial = f"{serial_row}-{i + 1}"
            else:
                serial = f"AUTO-{item.id_item}-{timestamp}-{activos_creados + i + 1}"

            res = await db.execute(select(Activo).where(Activo.serial == serial))
            if res.scalars().first():
                continue

            observaciones = f"Desmontado: {nombre}"
            if obs:
                observaciones += f" | {obs}"
            if activo_fijo:
                observaciones += f" | AF: {activo_fijo}"

            activo = Activo(
                id_item=item.id_item,
                serial=serial,
                estado_actual="LABORATORIO",
                condicion_fisica=condicion,
                ubicacion_fisica=ubicacion or "LABORATORIO",
                activo_fijo_securitas=activo_fijo,
                area_asignada="LABORATORIO",
                id_proyecto_actual=proyecto_id,
                id_cliente_actual=cliente_id,
                observaciones=observaciones,
            )
            db.add(activo)
            activos_creados += 1
            seriales_creados.append(serial)

        if seriales_creados:
            origen_str = "IMPORTACION_DESMONTES"
            if proyecto_id:
                origen_str = f"PROYECTO: {proyecto_id}"
            elif cliente_id:
                origen_str = f"CLIENTE: {cliente_id}"
            movimiento = MovimientoInventario(
                id_item=item.id_item,
                tipo_movimiento="INGRESO_DESMONTE",
                cantidad=cantidad,
                origen=origen_str,
                destino="LABORATORIO",
                notes=f"Importado desde plantilla de desmontes - Seriales: {', '.join(seriales_creados)}",
            )
            db.add(movimiento)

        await db.flush()

    return {
        "items": items_creados,
        "activos": activos_creados,
        "clientes": clientes_creados,
        "proyectos": proyectos_creados,
    }


async def _procesar_clientes(xl: pd.ExcelFile, db: AsyncSession):
    """
    Procesa Plantilla_Clientes.xlsx (hoja CLIENTES).
    """
    creados = 0
    actualizados = 0

    sheet_name = next((s for s in xl.sheet_names if "CLIENTE" in str(s).upper()), None)
    if not sheet_name:
        logger.warning("Plantilla de clientes sin hoja 'CLIENTES'.")
        return {"clientes": 0, "actualizados": 0}

    df = xl.parse(sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    TIPO_MAP = {
        "CORPORATIVO": "CORPORATIVO",
        "CORPORATIVA": "CORPORATIVO",
        "INTERNO": "INTERNO",
        "GENERAL": "GENERAL",
        "PUBLICO": "GENERAL",
    }

    for _, row in df.iterrows():
        nombre = _clean(row.get("NOMBRE"))
        if not nombre:
            continue

        nit = _clean(row.get("NIT"))
        if nit and nit.upper() in ("NAN", "SIN NIT"):
            nit = None

        result = await db.execute(
            select(Cliente).where(
                or_(
                    Cliente.nombre == nombre,
                    Cliente.nit == nit,
                )
                if nit
                else (Cliente.nombre == nombre)
            )
        )
        existing = result.scalars().first()
        if existing:
            actualizados += 1
            continue

        tipo_raw = _clean(row.get("TIPO_CLIENTE"))
        tipo = TIPO_MAP.get((tipo_raw or "CORPORATIVO").upper(), "CORPORATIVO")

        cliente = Cliente(
            nombre=nombre,
            nit=nit,
            contacto=_clean(row.get("CONTACTO")),
            email_contacto=_clean(row.get("EMAIL")) or _clean(row.get("CORREO")),
            telefono=_clean(row.get("TELEFONO")),
            direccion=_clean(row.get("DIRECCION")),
            ciudad=_clean(row.get("CIUDAD")),
            departamento=_clean(row.get("DEPARTAMENTO")),
            tipo_cliente=tipo,
            ceco_asociado=_clean(row.get("CECO")) or _clean(row.get("CECO_ASOCIADO")),
        )
        db.add(cliente)
        creados += 1

    await db.flush()
    return {"clientes": creados, "actualizados": actualizados}


async def _procesar_proveedores(xl: pd.ExcelFile, db: AsyncSession):
    """
    Procesa Plantilla_Proveedores.xlsx (hoja PROVEEDORES).
    """
    creados = 0
    actualizados = 0

    sheet_name = next(
        (s for s in xl.sheet_names if "PROVEEDOR" in str(s).upper()), None
    )
    if not sheet_name:
        logger.warning("Plantilla de proveedores sin hoja 'PROVEEDORES'.")
        return {"proveedores": 0, "actualizados": 0}

    df = xl.parse(sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    CAT_MAP = {
        "FABRICANTE": "FABRICANTE",
        "DISTRIBUIDOR": "DISTRIBUIDOR",
        "DISTRIBUIDORA": "DISTRIBUIDOR",
        "SERVICIO TECNICO": "SERVICIO_TECNICO",
        "SERVICIO_TECNICO": "SERVICIO_TECNICO",
        "LOGISTICA": "LOGISTICA",
        "LOGÍSTICA": "LOGISTICA",
    }

    for _, row in df.iterrows():
        nombre = _clean(row.get("NOMBRE"))
        if not nombre:
            continue

        nit = _clean(row.get("NIT"))
        if nit and nit.upper() in ("NAN", "SIN NIT"):
            nit = None

        result = await db.execute(
            select(Proveedor).where(
                or_(
                    Proveedor.nombre == nombre,
                    Proveedor.nit == nit,
                )
                if nit
                else (Proveedor.nombre == nombre)
            )
        )
        existing = result.scalars().first()
        if existing:
            actualizados += 1
            continue

        cat_raw = _clean(row.get("CATEGORIA"))
        categoria = CAT_MAP.get((cat_raw or "DISTRIBUIDOR").upper(), "DISTRIBUIDOR")

        proveedor = Proveedor(
            nombre=nombre,
            nit=nit,
            contacto=_clean(row.get("CONTACTO")),
            telefono=_clean(row.get("TELEFONO")),
            email=_clean(row.get("EMAIL")) or _clean(row.get("CORREO")),
            direccion=_clean(row.get("DIRECCION")),
            ciudad=_clean(row.get("CIUDAD")),
            dias_credito=int(_to_float(row.get("DIAS_CREDITO", 30) if "DIAS_CREDITO" in df.columns else 30, 30)),
            categoria=categoria,
        )
        db.add(proveedor)
        creados += 1

    await db.flush()
    return {"proveedores": creados, "actualizados": actualizados}


async def _procesar_proyectos(xl: pd.ExcelFile, db: AsyncSession):
    """
    Procesa Plantilla_Proyectos.xlsx (hoja PROYECTOS).
    """
    creados = 0
    actualizados = 0
    clientes_creados = 0

    sheet_name = next(
        (s for s in xl.sheet_names if "PROYECTO" in str(s).upper()), None
    )
    if not sheet_name:
        logger.warning("Plantilla de proyectos sin hoja 'PROYECTOS'.")
        return {"proyectos": 0, "actualizados": 0}

    df = xl.parse(sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    ESTADO_MAP = {
        "ACTIVO": "ACTIVO",
        "FINALIZADO": "FINALIZADO",
        "TERMINADO": "FINALIZADO",
        "CERRADO": "FINALIZADO",
        "PAUSADO": "PAUSADO",
        "SUSPENDIDO": "PAUSADO",
    }

    for _, row in df.iterrows():
        nombre = _clean(row.get("NOMBRE_PROYECTO")) or _clean(row.get("NOMBRE"))
        if not nombre:
            continue

        cliente_nombre = _clean(row.get("CLIENTE"))
        id_cliente = None
        if cliente_nombre:
            id_cliente = await _crear_cliente_si_no_existe(db, cliente_nombre)
            if id_cliente:
                clientes_creados += 1

        result = await db.execute(
            select(Proyecto).where(
                Proyecto.nombre_proyecto == nombre,
                Proyecto.id_cliente == id_cliente,
            )
        )
        existing = result.scalars().first()
        if existing:
            actualizados += 1
            continue

        estado_raw = _clean(row.get("ESTADO"))
        estado = ESTADO_MAP.get((estado_raw or "ACTIVO").upper(), "ACTIVO")

        proyecto = Proyecto(
            id_cliente=id_cliente,
            nombre_proyecto=nombre,
            descripcion=_clean(row.get("DESCRIPCION")),
            ubicacion=_clean(row.get("DIRECCION")) or _clean(row.get("CIUDAD")),
            estado=estado,
            fecha_inicio=_to_date(row.get("FECHA_INICIO")),
            fecha_fin_estimada=_to_date(row.get("FECHA_FIN_ESTIMADA")),
        )
        db.add(proyecto)
        creados += 1

    await db.flush()
    return {"proyectos": creados, "actualizados": actualizados, "clientes_creados": clientes_creados}


# ============================================================
# Endpoint principal: detecta el archivo y llama al procesador correcto
# ============================================================

ARCHIVO_TIPO = {
    "inventario_laboratorio": _procesar_inventario_laboratorio,
    "formato_inventario_clientes": _procesar_inventario_clientes,
    "asignacion_numero": _procesar_garantias,
    "desmontes": _procesar_desmontes,
    "clientes": _procesar_clientes,
    "proveedores": _procesar_proveedores,
    "proyectos": _procesar_proyectos,
}


def _detectar_tipo(filename: str) -> str | None:
    name = filename.lower().replace(" ", "_").replace("-", "_")
    name = (
        name.replace("á", "a")
        .replace("é", "e")
        .replace("ó", "o")
        .replace("ú", "u")
    )
    if "inventario_laboratorio" in name:
        return "inventario_laboratorio"
    if "inventario_clientes" in name or "formato_inventario" in name or "clientes_corporativos" in name:
        return "formato_inventario_clientes"
    if "asignacion" in name or "garantia" in name or "numero_de_caso" in name:
        return "asignacion_numero"
    if "desmonte" in name:
        return "desmontes"
    if "cliente" in name:
        return "clientes"
    if "proveedor" in name:
        return "proveedores"
    if "proyecto" in name:
        return "proyectos"
    return None


@router.get("/templates/{module}")
async def descargar_plantilla(module: str):
    """
    Descarga una plantilla Excel (.xlsx) para carga masiva de datos.
    Modulos: inventario_laboratorio, inventario_clientes, garantias, clientes, proveedores, desmontes, proyectos
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import FileResponse
    import tempfile

    TEMPLATES = {
        "inventario_laboratorio": {
            "filename": "Plantilla_Inventario_Laboratorio.xlsx",
            "sheets": {
                "INVENTARIO HERRAMIENTA": [
                    "MARCA", "DESCRIPCION DEL EQUIPO", "MODELO", "CANTIDAD", "UBICACIÓN", "ESTADO"
                ],
                "ESTANTES ACTUALIZADO": [
                    "CODIGO", "MARCA", "DESCRIPCION DEL EQUIPO", "MODELO", "CANTIDAD", "UBICACIÓN", "ACTIVO FIJO", "ESTADO", "OBSERVACIONES"
                ],
                "STOCK ELKIN": [
                    "CODIGO", "MARCA", "DESCRIPCION DEL EQUIPO", "MODELO", "CANTIDAD", "UBICACIÓN", "ACTIVO FIJO", "ESTADO"
                ],
                "ALSEA": [
                    "MARCA", "DESCRIPCION DEL EQUIPO", "MODELO", "CANTIDAD", "UBICACIÓN", "ACTIVO FIJO", "ESTADO", "TIENDA"
                ],
            }
        },
        "inventario_clientes": {
            "filename": "Plantilla_Inventario_Clientes.xlsx",
            "sheets": {
                "Inventario Consolidado": [
                    "SES", "Comercial", "Sistema", "Corporativo", "SECURITAS", "Ceco Compra",
                    "Codigo", "Item", "Marca", "Referencia", "Cantidad en Almacén", "Stock",
                    "Punto de recompra"
                ],
            }
        },
        "garantias": {
            "filename": "Plantilla_Garantias.xlsx",
            "sheets": {
                "GARANTIAS": [
                    "NUMERO DE CASO", "FECHA DE SOLICTUD", "CLIENTE", "PROYECTO",
                    "DESCRIPCIÓN DE EQUIPO", "REFERENCIA DE EQUIPO", "SERIAL",
                    "FALLA", "OBSERVACIÓNES DE EQUIPO (IP, CONTRASEÑAS)",
                    "APLICA GARANTA", "PROVEEDOR", "NUMERO DE FACTURA",
                    "RMA DEL PROVEEDOR", "FECHA DE ENVIO A PROVEDOR",
                    "COMENTARIOS DEL PROCESO", "ESTADO DEL EQUIPO",
                    "FECHA DE RECIBIDO DE EQUIPO", "OBSERVACION FINAL DEL PROCESO",
                    "AREA", "ESTADO DEL CASO"
                ],
                "STOCK MONITOREO": [
                    "NUMERO DE CASO", "DESCRIPCIÓN DE EQUIPO", "REFERENCIA DE EQUIPO", "PROYECTO NUEVO"
                ],
                "STOCK MANTENIMIENTO": [
                    "NUMERO DE CASO", "DESCRIPCIÓN DE EQUIPO", "REFERENCIA DE EQUIPO", "PROYECTO NUEVO", "ENTREGADO A"
                ],
                "STOCK INSTALACION": [
                    "NUMERO DE CASO", "DESCRIPCIÓN DE EQUIPO", "REFERENCIA DE EQUIPO", "PROYECTO NUEVO", "ENTREGADO A"
                ],
                "STOCK SOLUCIONES": [
                    "NUMERO DE CASO", "DESCRIPCIÓN DE EQUIPO", "REFERENCIA DE EQUIPO", "PROYECTO NUEVO"
                ],
            }
        },
        "clientes": {
            "filename": "Plantilla_Clientes.xlsx",
            "sheets": {
                "CLIENTES": [
                    "NOMBRE", "NIT", "CONTACTO", "EMAIL", "TELEFONO",
                    "DIRECCION", "CIUDAD", "DEPARTAMENTO", "TIPO_CLIENTE", "CECO"
                ]
            }
        },
        "proveedores": {
            "filename": "Plantilla_Proveedores.xlsx",
            "sheets": {
                "PROVEEDORES": [
                    "NOMBRE", "NIT", "CONTACTO", "TELEFONO", "EMAIL",
                    "DIRECCION", "CIUDAD", "DIAS_CREDITO", "CATEGORIA"
                ]
            }
        },
        "desmontes": {
            "filename": "Plantilla_Desmontes.xlsx",
            "sheets": {
                "DESMONTES": [
                    "MARCA", "DESCRIPCION DEL EQUIPO", "MODELO", "CANTIDAD",
                    "UBICACIÓN", "SERIAL", "ACTIVO FIJO", "ESTADO", "OBSERVACIONES",
                    "CLIENTE", "PROYECTO"
                ]
            }
        },
        "proyectos": {
            "filename": "Plantilla_Proyectos.xlsx",
            "sheets": {
                "PROYECTOS": [
                    "NOMBRE_PROYECTO", "DESCRIPCION", "CLIENTE",
                    "FECHA_INICIO", "FECHA_FIN_ESTIMADA", "ESTADO",
                    "DIRECCION", "CIUDAD", "RESPONSABLE"
                ]
            }
        },
    }

    if module not in TEMPLATES:
        raise HTTPException(status_code=404, detail=f"Plantilla '{module}' no disponible. Opciones: {', '.join(TEMPLATES.keys())}")

    tpl = TEMPLATES[module]
    wb = openpyxl.Workbook()
    ws_default = wb.active
    if ws_default is not None:
        wb.remove(ws_default)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for sheet_name, columns in tpl["sheets"].items():
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 4, 18)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=tpl["filename"],
        headers={"Content-Disposition": f'attachment; filename="{tpl["filename"]}"'}
    )


@router.post("/excel")
async def import_excel(
    file: UploadFile = File(...),
    id_proyecto: Optional[int] = None,
    id_cliente: Optional[int] = None,
    modo_stock: str = Form("sumar"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Importa cualquiera de los tres Excel del sistema.
    modo_stock: "sumar" (suma stock del Excel al existente) o "reemplazar" (pisa con el valor del Excel).
    """
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsx")

    MAX_IMPORT_SIZE = 30 * 1024 * 1024  # 30 MB
    content_length = file.size
    if content_length and content_length > MAX_IMPORT_SIZE:
        raise HTTPException(
            status_code=413,
            detail="El archivo excede el tamaño máximo permitido de 30 MB.",
        )
    contents = await file.read(MAX_IMPORT_SIZE + 1)
    if len(contents) > MAX_IMPORT_SIZE:
        raise HTTPException(
            status_code=413,
            detail="El archivo excede el tamaño máximo permitido de 30 MB.",
        )

    if modo_stock not in ("sumar", "reemplazar"):
        raise HTTPException(
            status_code=400,
            detail="modo_stock debe ser 'sumar' o 'reemplazar'.",
        )

    tipo = _detectar_tipo(file.filename or "")
    if not tipo:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Archivo '{file.filename}' no reconocido. "
                "Sube uno de: Inventario_laboratorio.xlsx, "
                "Formato_Inventario_Clientes_*.xlsx, "
                "ASIGNACION_NUMERO_DE_CASO_*.xlsx, "
                "Plantilla_Desmontes.xlsx, Plantilla_Clientes.xlsx, "
                "Plantilla_Proveedores.xlsx o Plantilla_Proyectos.xlsx"
            ),
        )

    try:
        xl = pd.ExcelFile(io.BytesIO(contents))
    except Exception as e:
        logger.warning(f"No se pudo leer el Excel '{file.filename}': {e}")
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo como Excel válido.")

    procesador = ARCHIVO_TIPO[tipo]
    try:
        if tipo in ("inventario_laboratorio", "desmontes"):
            resultado = await procesador(
                xl, db, id_proyecto=id_proyecto, id_cliente=id_cliente,
                modo_stock=modo_stock,
            )
        else:
            resultado = await procesador(xl, db, modo_stock=modo_stock)

        await db.commit()
        try:
            from app.crud.crud_alerts import evaluar_alertas
            await evaluar_alertas(db)
        except Exception:
            pass
        await create_audit_log(
            db,
            getattr(current_user, "id_usuario", 0),
            "importacion_datos",
            "CREATE",
            nuevo={
                "archivo": file.filename,
                "tipo_detectado": tipo,
                "modo_stock": modo_stock,
                "resultado": resultado,
            },
        )
        return {
            "mensaje": "Importación completada exitosamente.",
            "archivo": file.filename,
            "tipo_detectado": tipo,
            "modo_stock": modo_stock,
            "resultado": resultado,
        }
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error importando '{file.filename}': {e}", exc_info=True)
        if isinstance(e, (ValueError, StopIteration, IndexError)) and (
            "header" in str(e) or "lines in file" in str(e)
        ):
            raise HTTPException(
                status_code=400,
                detail="La plantilla no contiene filas de datos. Descargue la plantilla, llénela y súbala nuevamente.",
            )
        raise HTTPException(status_code=500, detail="Error durante la importación. Consulte los logs.")


# Endpoint legacy (mantiene compatibilidad con frontend existente)
@router.post("/full-system")
async def import_full_system(
    file: UploadFile = File(...),
    id_proyecto: Optional[int] = None,
    id_cliente: Optional[int] = None,
    modo_stock: str = Form("sumar"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Alias del endpoint /excel para compatibilidad con inventory.ts existente."""
    return await import_excel(
        file=file,
        id_proyecto=id_proyecto,
        id_cliente=id_cliente,
        modo_stock=modo_stock,
        db=db,
        current_user=current_user,
    )
