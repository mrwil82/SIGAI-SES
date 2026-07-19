import os
from io import BytesIO
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from xml.sax.saxutils import escape
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast, List, Dict, Any, Generator, Iterable
import xlsxwriter


# ── PALETA (tema oscuro de la app) ──

COLOR_DARK_HEADER   = "#0A0F0D"
COLOR_EMERALD       = "#00C26A"
COLOR_EMERALD_DEEP  = "#071F14"
COLOR_EMERALD_MUTED = "#0D3D26"
COLOR_SURFACE       = "#101820"
COLOR_CARD          = "#141F1A"
COLOR_LIGHTGRAY     = "#F5F6F7"
COLOR_BG_META       = "#F8F9FA"
COLOR_BORDER        = "#DDE1E5"
COLOR_TEXT_DARK     = "#1A1D1A"
COLOR_TEXT_MUTED    = "#5A7A65"
COLOR_WHITE         = "#FFFFFF"

ICON_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "static", "icon-app.png"
)

MODULE_TITLES = {
    "alerts"     : "REPORTE DE ALERTAS",
    "inventory"  : "REPORTE DE INVENTARIO",
    "guarantees" : "REPORTE DE GARANTÍAS",
    "clientes"   : "REPORTE DE CLIENTES",
    "projects"   : "REPORTE DE PROYECTOS",
    "actas"      : "REPORTE DE ACTAS DE ENTREGA",
    "acta"       : "ACTA DE ENTREGA DE HERRAMIENTA",
    "default"    : "REPORTE SIGAI-SES",
}


def get_module_title(module: str) -> str:
    return MODULE_TITLES.get(module.lower(), MODULE_TITLES["default"])


def safe_text(value: object) -> str:
    text = str(value) if value is not None else ""
    return escape(text).replace("\n", "<br/>")


def _build_logo_cell(
    sty_brand: ParagraphStyle,
    icon_width: float = 0.55,
    icon_height: float = 0.55,
    col_width: float = 1.6,
) -> Table:
    rows = []
    if os.path.exists(ICON_PATH):
        icon = RLImage(ICON_PATH, width=icon_width * inch, height=icon_height * inch)
        rows.append([icon])
    rows.append([Paragraph("<b>SIGAI-SES</b>", sty_brand)])
    tbl = Table(rows, colWidths=[col_width * inch])
    tbl.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    return tbl


def _meta_column_spans(n_data_cols: int) -> List[int]:
    if n_data_cols <= 0:
        return [1, 1, 1, 1]
    if n_data_cols <= 4:
        return [1] * n_data_cols + [0] * (4 - n_data_cols)
    base = n_data_cols // 4
    extra = n_data_cols % 4
    return [base + (1 if i < extra else 0) for i in range(4)]


# ── PDF: Encabezado institucional ──

def build_pdf_header(module: str, meta: dict | None = None) -> list:
    styles = getSampleStyleSheet()
    title = get_module_title(module)
    meta = meta or {}

    sty_title = ParagraphStyle(
        "SecTitle", parent=styles["Normal"],
        fontSize=16, fontName="Helvetica-Bold",
        alignment=0, textColor=colors.white, leading=20,
    )
    sty_brand = ParagraphStyle(
        "Brand", parent=styles["Normal"],
        fontSize=8, fontName="Helvetica-Bold",
        alignment=1, textColor=colors.HexColor(COLOR_EMERALD), leading=10,
    )
    sty_meta_label = ParagraphStyle(
        "MetaLabel", parent=styles["Normal"],
        fontSize=7, fontName="Helvetica-Bold",
        textColor=colors.HexColor(COLOR_TEXT_MUTED), alignment=1,
    )
    sty_meta_value = ParagraphStyle(
        "MetaValue", parent=styles["Normal"],
        fontSize=8, fontName="Helvetica",
        textColor=colors.HexColor(COLOR_TEXT_DARK), alignment=1,
    )

    # ── Fila 1: Icono + SIGAI-SES + Título ──
    logo_cell = _build_logo_cell(sty_brand, icon_width=0.55, icon_height=0.55, col_width=1.6)

    t1 = Table(
        [[logo_cell, Paragraph(f"<b>{title}</b>", sty_title)]],
        colWidths=[1.6 * inch, 6.1 * inch],
    )
    t1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(COLOR_DARK_HEADER)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("NOSPLIT", (0, 0), (-1, -1)),
    ]))

    # Línea decorativa esmeralda
    t_accent = Table([[""]], colWidths=[7.7 * inch])
    t_accent.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(COLOR_EMERALD)),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))

    # ── Metadatos ──
    labels = ["MÓDULO", "FECHA DE GENERACIÓN", "CARGO", "GENERADO POR"]
    values = [
        module.upper(),
        meta.get("fecha", datetime.now().strftime("%d/%m/%Y %H:%M")),
        meta.get("cargo", "—").upper(),
        meta.get("generado_por", "SISTEMA").upper(),
    ]

    t2 = Table(
        [
            [Paragraph(l, sty_meta_label) for l in labels],
            [Paragraph(v, sty_meta_value) for v in values],
        ],
        colWidths=[1.925 * inch] * 4,
    )
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(COLOR_BG_META)),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(COLOR_BORDER)),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor(COLOR_BORDER)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
        ("TOPPADDING", (0, 1), (-1, 1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))

    return [t1, t_accent, Spacer(1, 6), t2, Spacer(1, 12)]


# ── EXCEL (Streaming - XlsxWriter) ──

def stream_excel_large(
    data_generator: Iterable[Dict[str, Any]],
    module: str,
    meta: dict | None,
    columns: List[str],
) -> bytes:
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'constant_memory': True, 'in_memory': True})
    worksheet = workbook.add_worksheet(module[:31])
    n_cols = max(len(columns), 1)

    fmt_header = workbook.add_format({
        'bold': True, 'bg_color': COLOR_DARK_HEADER, 'font_color': 'white',
        'border': 0, 'align': 'center', 'valign': 'vcenter',
        'font_size': 9, 'font_name': 'Calibri',
    })
    fmt_cell = workbook.add_format({
        'border': 1, 'border_color': COLOR_BORDER, 'font_size': 9,
        'align': 'left', 'valign': 'vcenter', 'font_name': 'Calibri',
    })
    fmt_cell_alt = workbook.add_format({
        'border': 1, 'border_color': COLOR_BORDER, 'font_size': 9,
        'align': 'left', 'valign': 'vcenter', 'font_name': 'Calibri',
        'bg_color': COLOR_LIGHTGRAY,
    })
    fmt_title = workbook.add_format({
        'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter',
        'font_color': 'white', 'bg_color': COLOR_DARK_HEADER, 'font_name': 'Calibri',
    })
    fmt_brand = workbook.add_format({
        'bold': True, 'font_size': 8, 'align': 'center', 'valign': 'vcenter',
        'font_color': COLOR_EMERALD, 'bg_color': COLOR_DARK_HEADER, 'font_name': 'Calibri',
    })
    fmt_meta_lbl = workbook.add_format({
        'bold': True, 'bg_color': COLOR_BG_META, 'font_color': COLOR_TEXT_MUTED,
        'border': 1, 'border_color': COLOR_BORDER,
        'align': 'center', 'valign': 'vcenter', 'font_size': 8, 'font_name': 'Calibri',
    })
    fmt_meta_val = workbook.add_format({
        'bg_color': COLOR_BG_META, 'border': 1, 'border_color': COLOR_BORDER,
        'align': 'center', 'valign': 'vcenter', 'font_size': 9, 'font_name': 'Calibri',
    })

    # ── Fila 1: Icono + Título ──
    worksheet.set_row(0, 48)
    last_data_col = n_cols - 1
    worksheet.merge_range(0, 1, 0, last_data_col, get_module_title(module), fmt_title)
    worksheet.write(0, 0, "SIGAI-SES", fmt_brand)

    if os.path.exists(ICON_PATH):
        worksheet.insert_image(
            0, 0, ICON_PATH,
            {'x_scale': 0.035, 'y_scale': 0.035, 'x_offset': 4, 'y_offset': 2}
        )

    # ── Filas 2-3: Metadatos (ocupan todo el ancho) ──
    meta = meta or {}
    labels = ["MÓDULO", "FECHA DE GENERACIÓN", "CARGO", "GENERADO POR"]
    values = [
        module.upper(),
        meta.get("fecha", datetime.now().strftime("%d/%m/%Y %H:%M")),
        meta.get("cargo", "N/A").upper(),
        meta.get("generado_por", "SISTEMA").upper(),
    ]
    spans = _meta_column_spans(n_cols)

    col = 0
    for i, (lbl, val) in enumerate(zip(labels, values)):
        span = spans[i]
        if span == 0:
            break
        if span > 1:
            worksheet.merge_range(1, col, 1, col + span - 1, lbl, fmt_meta_lbl)
            worksheet.merge_range(2, col, 2, col + span - 1, val, fmt_meta_val)
        else:
            worksheet.write(1, col, lbl, fmt_meta_lbl)
            worksheet.write(2, col, val, fmt_meta_val)
        col += span

    # Anchos de columna calculados del contenido real
    all_data = list(data_generator) if not isinstance(data_generator, list) else data_generator
    for ci in range(n_cols):
        col_name = str(columns[ci]) if ci < len(columns) else ""
        max_len = len(col_name) + 2
        for row in all_data:
            val = row.get(col_name, "")
            cell_len = len(str(val))
            if cell_len > max_len:
                max_len = cell_len
        width = min(max(max_len + 3, 10), 50)
        worksheet.set_column(ci, ci, width)
    data_generator = all_data

    worksheet.set_row(3, 4)

    # ── Encabezados de tabla ──
    start_row = 4
    worksheet.set_row(start_row, 22)
    for ci, col_name in enumerate(columns):
        worksheet.write(start_row, ci, col_name.upper(), fmt_header)

    # ── Datos ──
    current_row = start_row + 1
    row_count = 0
    for row_dict in data_generator:
        fmt = fmt_cell_alt if row_count % 2 == 0 else fmt_cell
        for ci, col_name in enumerate(columns):
            val = row_dict.get(col_name, "")
            if isinstance(val, (int, float)):
                worksheet.write_number(current_row, ci, val, fmt)
            else:
                worksheet.write(current_row, ci, str(val) if val is not None else "", fmt)
        current_row += 1
        row_count += 1

    workbook.close()
    output.seek(0)
    return output.getvalue()


# ── PDF: Exportación de datos ──

def export_to_pdf(data: list, module: str, meta: dict | None = None) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=0.5 * inch, leftMargin=0.5 * inch,
        topMargin=0.4 * inch, bottomMargin=0.5 * inch,
    )

    elements = build_pdf_header(module, meta)
    styles = getSampleStyleSheet()

    if data:
        df = pd.DataFrame(data)
        cell_sty = ParagraphStyle(
            "C", parent=styles["Normal"], fontSize=7.5, leading=11,
            fontName="Helvetica",
        )
        hdr_sty = ParagraphStyle(
            "H", parent=styles["Normal"], fontSize=8,
            fontName="Helvetica-Bold", textColor=colors.white, alignment=1,
        )

        headers = [str(c) for c in df.columns]
        table_data = [[Paragraph(h, hdr_sty) for h in headers]]

        for row in df.astype(str).values:
            row_cells = [
                Paragraph(str(v) if v and v != "nan" else "", cell_sty)
                for v in row
            ]
            table_data.append(row_cells)

        avail = doc.pagesize[0] - 1.0 * inch
        n_cols = len(df.columns)
        col_w = avail / n_cols

        tbl = Table(table_data, colWidths=[col_w] * n_cols, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(COLOR_DARK_HEADER)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(COLOR_BORDER)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]
        for ri in range(1, len(table_data)):
            if ri % 2 == 0:
                style_cmds.append(
                    ("BACKGROUND", (0, ri), (-1, ri), colors.HexColor(COLOR_LIGHTGRAY))
                )
        tbl.setStyle(TableStyle(style_cmds))
        elements.append(tbl)
    else:
        elements.append(Paragraph("Sin datos disponibles.", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ── EXCEL (OpenPyXL - datos pequeños) ──

def export_to_excel(data: list, module: str, meta: dict | None = None) -> BytesIO:
    output = BytesIO()
    df = pd.DataFrame(data) if data else pd.DataFrame()
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = module[:31]

    from openpyxl.drawing.image import Image as XLImage

    meta = meta or {}
    title = get_module_title(module)
    n_cols = len(df.columns) if not df.empty else 4
    labels = ["MÓDULO", "FECHA DE GENERACIÓN", "CARGO", "GENERADO POR"]
    values = [
        module.upper(),
        meta.get("fecha", datetime.now().strftime("%d/%m/%Y %H:%M")),
        meta.get("cargo", "N/A").upper(),
        meta.get("generado_por", "SISTEMA").upper(),
    ]

    hdr_fill = PatternFill("solid", fgColor=COLOR_DARK_HEADER.replace("#", ""))
    label_fill = PatternFill("solid", fgColor=COLOR_BG_META.replace("#", ""))
    alt_fill = PatternFill("solid", fgColor=COLOR_LIGHTGRAY.replace("#", ""))
    border_thin = Border(
        left=Side(style="thin", color=COLOR_BORDER.replace("#", "")),
        right=Side(style="thin", color=COLOR_BORDER.replace("#", "")),
        top=Side(style="thin", color=COLOR_BORDER.replace("#", "")),
        bottom=Side(style="thin", color=COLOR_BORDER.replace("#", "")),
    )

    # ── Fila 1: Título ──
    ws.row_dimensions[1].height = 42
    last_col = max(n_cols, 4)
    cast(Cell, ws["B1"]).value = title
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    cast(Cell, ws["B1"]).font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    cast(Cell, ws["B1"]).alignment = Alignment(horizontal="left", vertical="center")
    cast(Cell, ws["B1"]).fill = hdr_fill
    for col_idx in range(1, last_col + 1):
        c = cast(Cell, ws.cell(row=1, column=col_idx))
        c.fill = hdr_fill
        c.border = Border()

    if os.path.exists(ICON_PATH):
        try:
            img = XLImage(ICON_PATH)
            img.width = 28
            img.height = 28
            img.anchor = "A1"
            ws.add_image(img)
        except:
            cast(Cell, ws["A1"]).value = ""
    cast(Cell, ws["A1"]).value = "SIGAI-SES"
    cast(Cell, ws["A1"]).font = Font(bold=True, size=7, color=COLOR_EMERALD, name="Calibri")
    cast(Cell, ws["A1"]).fill = hdr_fill
    cast(Cell, ws["A1"]).alignment = Alignment(horizontal="center", vertical="bottom")

    # ── Metadatos (ocupan todo el ancho) ──
    ws.row_dimensions[2].height = 20
    spans = _meta_column_spans(n_cols)

    col = 1
    for i, (lbl, val) in enumerate(zip(labels, values)):
        span = spans[i]
        if span == 0:
            break
        if span > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + span - 1)
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + span - 1)
        c_lbl = cast(Cell, ws.cell(row=2, column=col))
        c_lbl.value = lbl
        c_lbl.fill = label_fill
        c_lbl.font = Font(bold=True, color=COLOR_TEXT_MUTED.replace("#", ""), size=8, name="Calibri")
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")

        c_val = cast(Cell, ws.cell(row=3, column=col))
        c_val.value = val
        c_val.fill = label_fill
        c_val.font = Font(size=9, name="Calibri")
        c_val.alignment = Alignment(horizontal="center", vertical="center")

        for c in range(col, col + span):
            ws.column_dimensions[get_column_letter(c)].width = 22
        col += span

    ws.row_dimensions[4].height = 6

    # ── Tabla de datos ──
    first_row = 5
    if not df.empty:
        ws.row_dimensions[first_row].height = 22
        for ci in range(1, n_cols + 1):
            cell_header = cast(Cell, ws.cell(row=first_row, column=ci))
            col_name = str(df.columns[ci - 1]).upper()
            cell_header.value = col_name
            cell_header.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
            cell_header.fill = hdr_fill
            cell_header.alignment = Alignment(horizontal="center", vertical="center")
            cell_header.border = border_thin
            estimated = max(len(col_name) + 3, 14)
            ws.column_dimensions[get_column_letter(ci)].width = min(estimated, 45)

        for ri, row in enumerate(df.values, first_row + 1):
            ws.row_dimensions[ri].height = 18
            for ci, val in enumerate(row, 1):
                c = cast(Cell, ws.cell(row=ri, column=ci))
                c.value = val if val is not None else ""
                c.font = Font(size=9, name="Calibri")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = border_thin
                if ri % 2 == 0:
                    c.fill = alt_fill

    wb.save(output)
    output.seek(0)
    return output


# ── ACTA DE ENTREGA (PDF) ──

def generate_pdf_acta(data: dict) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        topMargin=0.5 * inch, bottomMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()

    style_header = ParagraphStyle(
        'ActaHeader', parent=styles['Heading2'],
        alignment=1, fontSize=13, spaceAfter=6,
        textColor=colors.white, fontName='Helvetica-Bold',
    )
    style_section = ParagraphStyle(
        'SectionTitle', parent=styles['Heading2'],
        fontSize=10, spaceAfter=6, spaceBefore=4,
        textColor=colors.HexColor(COLOR_DARK_HEADER),
        fontName='Helvetica-Bold',
    )
    style_legal = ParagraphStyle(
        'LegalStyle', parent=styles['Normal'],
        fontSize=8, alignment=4, leading=11, fontName='Helvetica',
    )
    style_label = ParagraphStyle(
        'LabelStyle', parent=styles['Normal'],
        fontSize=8, fontName='Helvetica-Bold',
        textColor=colors.HexColor(COLOR_TEXT_MUTED),
    )
    style_value = ParagraphStyle(
        'ValueStyle', parent=styles['Normal'],
        fontSize=9, fontName='Helvetica',
        textColor=colors.HexColor(COLOR_TEXT_DARK),
    )
    table_cell_style = ParagraphStyle(
        'TableCell', parent=styles['Normal'],
        fontSize=7.5, leading=10, fontName='Helvetica',
    )
    header_style = ParagraphStyle(
        'TableHeader', parent=styles['Normal'],
        fontSize=7.5, leading=10, fontName='Helvetica-Bold',
        textColor=colors.white,
    )
    sty_brand = ParagraphStyle(
        "Brand", parent=styles["Normal"],
        fontSize=7, fontName="Helvetica-Bold",
        alignment=1, textColor=colors.HexColor(COLOR_EMERALD), leading=9,
    )

    elements = []

    # ── Encabezado ──
    logo_cell = _build_logo_cell(sty_brand, icon_width=0.5, icon_height=0.5, col_width=1.6)

    header_data = [[logo_cell, Paragraph("<b>ACTA DE ENTREGA<br/>DE HERRAMIENTA</b>", style_header)]]
    t_header = Table(header_data, colWidths=[1.8 * inch, 5.4 * inch])
    t_header.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(COLOR_DARK_HEADER)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t_header)

    t_accent = Table([[""]], colWidths=[7.2 * inch])
    t_accent.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(COLOR_EMERALD)),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(t_accent)
    elements.append(Spacer(1, 12))

    # ── 1. Información del Técnico ──
    elements.append(Paragraph("1. INFORMACIÓN DEL TÉCNICO", style_section))

    tecnico_data = [
        [Paragraph("NOMBRE", style_label), Paragraph("FECHA DE ENTREGA", style_label)],
        [Paragraph(f"<b>{data.get('nombre_tecnico', '').upper() or '—'}</b>", style_value),
         Paragraph(data.get('fecha', datetime.now().strftime("%d/%m/%Y")), style_value)],
        [Paragraph("REGIONAL", style_label), Paragraph("", style_label)],
        [Paragraph(f"<b>{data.get('regional', 'BOGOTÁ').upper() or '—'}</b>", style_value),
         Paragraph("", style_value)],
    ]

    t_tecnico = Table(tecnico_data, colWidths=[3.6 * inch, 3.6 * inch])
    t_tecnico.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(COLOR_BORDER)),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_BG_META)),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor(COLOR_BG_META)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t_tecnico)
    elements.append(Spacer(1, 12))

    # ── 2. Herramienta y/o Equipos ──
    elements.append(Paragraph("2. HERRAMIENTA Y/O EQUIPOS", style_section))

    headers = ["ITEM", "DESCRIPCIÓN", "MARCA", "REFERENCIA", "SERIE", "CANT.", "OBSERVACIONES"]
    table_data = [[Paragraph(h, header_style) for h in headers]]

    raw_items = data.get('items', [])
    for i in range(17):
        if i < len(raw_items):
            item = raw_items[i]
            row = [
                str(i + 1),
                item.get('descripcion', ''),
                item.get('marca', ''),
                item.get('referencia', ''),
                item.get('serie', ''),
                str(item.get('cantidad', 1)),
                item.get('observaciones', ''),
            ]
        else:
            row = [str(i + 1), '', '', '', '', '', '']
        table_data.append([Paragraph(safe_text(c), table_cell_style) for c in row])

    t_items = Table(
        table_data,
        colWidths=[0.35 * inch, 2.4 * inch, 0.75 * inch, 0.85 * inch, 0.85 * inch, 0.4 * inch, 1.2 * inch],
        repeatRows=1,
    )
    style_items = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_DARK_HEADER)),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(COLOR_BORDER)),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]
    for ri in range(1, len(table_data)):
        if ri % 2 == 0:
            style_items.append(
                ('BACKGROUND', (0, ri), (-1, ri), colors.HexColor(COLOR_LIGHTGRAY))
            )
    t_items.setStyle(TableStyle(style_items))
    elements.append(t_items)
    elements.append(Spacer(1, 12))

    # ── 3. Observaciones ──
    elements.append(Paragraph("3. OBSERVACIONES", style_section))
    obs_text = data.get('observaciones_generales', '')
    elements.append(Paragraph(
        safe_text(obs_text) if obs_text else "—", styles['Normal']
    ))
    elements.append(Spacer(1, 15))

    # ── 4. Cláusula legal ──
    clausula = (
        "<b>AUTORIZACIÓN ESPECIAL DEL TRABAJADOR:</b> "
        "Autorizo expresamente a SECURITAS para que, por conducto de la dependencia "
        "correspondiente, descuente de mi salario los valores que se determinen como "
        "responsabilidad mía por concepto de daños, hurto, pérdida o extravío de los "
        "elementos aquí relacionados, de acuerdo con el procedimiento establecido en el "
        "reglamento interno de trabajo y las normas legales vigentes."
    )
    elements.append(Paragraph(clausula, style_legal))
    elements.append(Spacer(1, 25))

    # ── 5. Firmas ──
    firmas_data = [
        [
            Paragraph("<b>____________________________</b><br/>REPRESENTANTE", styles['Normal']),
            Paragraph("<b>____________________________</b><br/>TÉCNICO", styles['Normal']),
        ],
        [Spacer(1, 0.6 * inch), Spacer(1, 0.6 * inch)],
        [
            Paragraph("Nombre: ___________________<br/>CC: ___________________",
                      ParagraphStyle('F1', parent=styles['Normal'], fontSize=8, alignment=1)),
            Paragraph("Nombre: ___________________<br/>CC: ___________________",
                      ParagraphStyle('F2', parent=styles['Normal'], fontSize=8, alignment=1)),
        ],
    ]
    t_firmas = Table(firmas_data, colWidths=[3.5 * inch, 3.5 * inch])
    t_firmas.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor(COLOR_BORDER)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t_firmas)

    doc.build(elements)
    buffer.seek(0)
    return buffer
